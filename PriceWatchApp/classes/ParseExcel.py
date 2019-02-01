from openpyxl import load_workbook


class ParseExcel:

    def __init__(self, fileName):
        self.fileName = fileName
        self.wb = load_workbook(self.fileName)
        self.sheet_names = self.wb.sheetnames

    def getSheet(self, sheet_name):
        sheet = self.wb[sheet_name]
        return sheet

    # get all sheet names in workbook
    def getSheets(self):
        return self.sheet_names

    def getParsedExcelList(self, sheet):
        try:
            product_list = list()
            for row in sheet.rows:
                product_list.append((row[0].value, row[1].value, row[2].value, row[3].value, row[4].value, row[5].value))
            return product_list

        except Exception as e:
            print(e)
            return False


    def addToDatabase(self, sheet):
        try:
            for row in sheet.rows:
                insert_contact(row[0].value, row[1].value, row[2].value, row[3].value, row[4].value)
            return True

        except Exception as e:
            return e


    def selectAll(self):
        items = select_all()
        for i in items:
            print(i)



